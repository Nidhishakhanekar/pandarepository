import openpyxl


def Readdata(File, sheet, col, rownumber):
    workbook = openpyxl.load_workbook(File)
    sheetName = workbook[sheet]
    return sheetName.cell(row=rownumber, column=col).value


def Writedata(File, sheet, col, rownumber, data):
    workbook = openpyxl.load_workbook(File)
    sheetName = workbook[sheet]
    sheetName.cell(row=rownumber, column=col).value = data
    workbook.save(File)


